"""
Southport Agencies — Excel Data Ingestion Pipeline
===================================================
Parses all SOUTHPORT-LINEUP and FGIS Report Excel files from the Southport
Agencies data feed into DuckDB tables in the platform analytics database.

Tables created/replaced:
  grain_southport_lineup     — one row per vessel/terminal per report date
  grain_miss_elevator_ref    — Mississippi River elevator reference (from LINEUP header)
  grain_fgis_certs           — FGIS grain export certificates (detailed per-shipment)
  grain_ingest_log           — tracks which files have been processed

Usage:
  python southport_ingest.py                   # ingest all new files
  python southport_ingest.py --lineup          # LINEUP files only
  python southport_ingest.py --fgis            # FGIS files only
  python southport_ingest.py --force           # reprocess all files
  python southport_ingest.py --summary         # print table row counts
"""

from __future__ import annotations

import re
import sys
import json
import argparse
import datetime
from pathlib import Path
from typing import Optional

import duckdb
import openpyxl

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).resolve().parent
MODULE_ROOT  = SCRIPT_DIR.parent
PROJECT_ROOT = MODULE_ROOT.parent.parent   # project_master_reporting/

SOUTHPORT_DIR = Path("/Users/wsd/Projects/southport-agencies/exposed-files/excel")
DB_PATH       = PROJECT_ROOT / "data" / "analytics.duckdb"

# ── Commodity code → full name ───────────────────────────────────────────────
COMMODITY_MAP = {
    "YSB":    "Yellow Soybeans",
    "CORN":   "Corn",
    "WHT":    "Wheat",
    "SBM":    "Soybean Meal",
    "DDGS":   "Distillers Dried Grains (DDGS)",
    "SORG":   "Sorghum",
    "SRW":    "Soft Red Winter Wheat",
    "HRS":    "Hard Red Spring Wheat",
    "HRW":    "Hard Red Winter Wheat",
    "SWW":    "Soft White Wheat",
    "BARLEY": "Barley",
    "RICE":   "Rice",
    "RVT":    None,   # "Reserved — to be determined"
}


def normalize_commodity(raw: Optional[str]) -> tuple[str | None, str | None]:
    """Return (commodity_code, commodity_name). Handles blends like 'CORN/WHT'."""
    if not raw or str(raw).strip().upper() in ("RVT", ""):
        return None, None
    code = str(raw).strip().upper()
    name = COMMODITY_MAP.get(code)
    if "/" in code:
        # blend — return as-is for code, build name from parts
        parts = [COMMODITY_MAP.get(p.strip(), p.strip()) for p in code.split("/")]
        name = " / ".join(p for p in parts if p)
    return code if code != "RVT" else None, name


def parse_mt(raw: Optional[str]) -> Optional[float]:
    """Convert MT string to numeric metric tons. '60K'→60000, '27.5K'→27500, RVT→None."""
    if raw is None:
        return None
    s = str(raw).strip().upper().replace(",", "")
    if s in ("RVT", "", "N/A", "NONE", "TBD"):
        return None
    # Remove trailing 'MT' or 'K MT'
    s = re.sub(r"\s*MT$", "", s).strip()
    if s.endswith("K"):
        try:
            return float(s[:-1]) * 1000
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_delay(delay_str: str) -> tuple[Optional[int], Optional[int]]:
    """Parse 'EXPECTED DELAYS: 4-6 DAYS' → (4, 6). 'NONE' → (0, 0)."""
    s = str(delay_str).upper()
    if "NONE" in s or "0 DAY" in s:
        return 0, 0
    m = re.search(r"(\d+)\s*[-–]\s*(\d+)", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"(\d+)\s*DAY", s)
    if m:
        v = int(m.group(1))
        return v, v
    return None, None


def parse_date_from_filename(filename: str) -> Optional[datetime.date]:
    """Extract report date from SOUTHPORT-LINEUP-MM.DD.YYYY.xlsx filenames."""
    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", filename)
    if m:
        try:
            return datetime.date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None


def parse_eta_date(status_str: str, report_date: datetime.date) -> Optional[datetime.date]:
    """Parse 'ETA 01/07' into a date. Infers year from report_date context."""
    m = re.search(r"ETA\s+(\d{1,2})/(\d{1,2})", str(status_str).upper())
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        year = report_date.year
        # If reported month < report month, vessel is in next year
        if month < report_date.month and report_date.month >= 11:
            year += 1
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return None
    return None


# ── MISS RIVER reference table (header block in each LINEUP file) ────────────

def parse_miss_elevator_ref(ws_rows: list[tuple], report_date: datetime.date) -> list[dict]:
    """Extract the elevator reference table from the MISS RIVER LINEUP header (rows 12-22)."""
    records = []
    in_ref = False
    for row in ws_rows:
        b = str(row[1] or "").strip()
        if "LOCATION" in b.upper() and "RIVER MILES" in b.upper():
            in_ref = True
            continue
        if not in_ref:
            continue
        # End of reference block: next terminal section starts when col A is non-empty string
        if row[0] and str(row[0]).strip() and str(row[0]).strip() not in (" ", ""):
            break
        # Reference rows: col B = mile (float), col C = elevator name
        if row[1] is not None and isinstance(row[1], (int, float)):
            records.append({
                "report_date":       str(report_date),
                "mile_ahp":          float(row[1]),
                "elevator_name":     str(row[2] or "").strip() if row[2] else None,
                "air_height_spouts": str(row[5] or "").strip() if len(row) > 5 else None,
                "gauge_reference":   str(row[6] or "").strip() if len(row) > 6 else None,
            })
    return records


# ── LINEUP sheet parser ───────────────────────────────────────────────────────

def _is_vessel_header(row: tuple, col_offset: int) -> bool:
    """Row is a VESSEL/ATA/STATUS/MT/COMMODITY/DESTINATION column header."""
    b = str(row[col_offset + 1] or "").strip().upper()
    return b in ("VESSEL:", "VESSEL")


def _is_delay_row(row: tuple, col_offset: int) -> bool:
    b = str(row[col_offset + 1] or "").strip().upper()
    return b.startswith("EXPECTED DELAYS")


def _is_none_loading(row: tuple, col_offset: int) -> bool:
    b = str(row[col_offset + 1] or "").strip().upper()
    return "NONE LOADING" in b or b == "NONE"


def _is_blank(row: tuple) -> bool:
    return not any(v is not None and str(v).strip() for v in row)


def _is_weather_section(row: tuple) -> bool:
    b = str(row[1] or "").strip().upper()
    return "WEATHER FORECAST" in b or "PILOT" in b or "HOLIDAY" in b


def _extract_terminal_abbrev(row: tuple, col_offset: int) -> Optional[str]:
    """Pull terminal abbreviation from col A (Miss/TX) or col B (PNW)."""
    v = row[col_offset]
    if v and str(v).strip():
        return str(v).strip()
    return None


def parse_lineup_sheet(ws, region: str, col_offset: int, report_date: datetime.date,
                       source_file: str) -> list[dict]:
    """
    Parse one LINEUP sheet (MISS RIVER, TEXAS, or PNW) into vessel records.

    col_offset:
      0 = MISS RIVER / TEXAS  (cols A-G carry terminal, vessel, ata, status, mt, commodity, dest)
      1 = PNW                 (cols B-H carry the same, col A is always None)
    """
    rows = list(ws.iter_rows(values_only=True))
    records = []

    current_terminal_abbrev: Optional[str] = None
    current_terminal_name: Optional[str] = None
    current_mile_ahp: Optional[float] = None
    in_vessel_block = False
    # Buffer for current terminal's vessel rows — backfill delays when EXPECTED DELAYS row found
    terminal_buffer: list[dict] = []

    def _flush_buffer(delay_min=None, delay_max=None):
        for rec in terminal_buffer:
            rec["delay_days_min"] = delay_min
            rec["delay_days_max"] = delay_max
        records.extend(terminal_buffer)
        terminal_buffer.clear()

    for row in rows:
        if _is_blank(row):
            continue
        if _is_weather_section(row):
            _flush_buffer()
            break

        # Detect vessel-header row (VESSEL:, ATA:, STATUS:, ...)
        if _is_vessel_header(row, col_offset):
            # Flush previous terminal buffer (no delay row found before next terminal)
            _flush_buffer()
            # New terminal block starts
            in_vessel_block = True
            # Terminal abbreviation comes from col A (or col B for PNW)
            abbrev = _extract_terminal_abbrev(row, col_offset)
            if abbrev:
                current_terminal_abbrev = abbrev
            continue

        if not in_vessel_block:
            # Look for terminal section info rows (mile marker lines)
            b = str(row[col_offset + 1] or "").strip()
            if "MILE" in b.upper() and "AHP" in b.upper():
                # Extract mile number: "MILE 61.5 AHP (6.4MBU)"
                m = re.search(r"MILE\s+([\d.]+)\s*AHP", b, re.IGNORECASE)
                if m:
                    current_mile_ahp = float(m.group(1))
            # Look for terminal name in section header lines
            # (e.g., "CHS MYRTLE GROVE-" or "COLUMBIA EXPORT – T5 – PORTLAND")
            elif b and b not in (" ", "") and not any(
                kw in b.upper() for kw in ("GRAIN LINEUP", "VESSEL", "LOCATION", "RIVER MILE", "ATA")
            ):
                # Could be terminal full name
                raw_name = b.rstrip("-– ").strip()
                if len(raw_name) > 2:
                    current_terminal_name = raw_name
            continue

        # ---- We are inside a vessel block ----

        if _is_delay_row(row, col_offset):
            b = str(row[col_offset + 1] or "")
            d_min, d_max = parse_delay(b)
            _flush_buffer(d_min, d_max)
            in_vessel_block = False
            continue

        if _is_none_loading(row, col_offset):
            # No vessels at this terminal — nothing to buffer
            continue

        # Vessel data row
        # col layout (adjusted for offset):
        # [offset+0]=terminal_abbrev, [offset+1]=vessel_name,
        # [offset+2]=ata_or_none,    [offset+3]=status_or_eta,
        # [offset+4]=mt,             [offset+5]=commodity, [offset+6]=destination
        try:
            r = row
            abbrev_v    = r[col_offset]
            vessel_name = str(r[col_offset + 1] or "").strip()
            ata_raw     = r[col_offset + 2]
            status_raw  = str(r[col_offset + 3] or "").strip() if len(r) > col_offset + 3 else ""
            mt_raw      = r[col_offset + 4] if len(r) > col_offset + 4 else None
            commodity_r = r[col_offset + 5] if len(r) > col_offset + 5 else None
            dest_raw    = r[col_offset + 6] if len(r) > col_offset + 6 else None
        except (IndexError, TypeError):
            continue

        # Skip non-vessel rows that slip through
        if not vessel_name or vessel_name.upper() in (
            "VESSEL:", "VESSEL", "NAME", "ATA:", "ATA"
        ):
            continue
        # Skip informational text rows
        if any(kw in vessel_name.upper() for kw in (
            "EXPECTED", "WEATHER", "PLEASE", "REFER", "PILOT", "HOLIDAY",
            "EXTENDED", "CLOSED", "NOTE:", "ALL ADM", "AFTER FILING"
        )):
            continue

        # Update terminal abbreviation if the column is populated on this row
        if abbrev_v and str(abbrev_v).strip():
            current_terminal_abbrev = str(abbrev_v).strip()

        # Determine status and dates
        ata_date: Optional[str] = None
        eta_date: Optional[str] = None
        status_type: Optional[str] = None

        if isinstance(ata_raw, datetime.datetime):
            ata_date = ata_raw.date().isoformat()
            st = status_raw.upper()
            if "LOADING" in st:
                status_type = "LOADING"
            elif "FILED" in st:
                status_type = "FILED"
            else:
                status_type = status_raw.upper() or "UNKNOWN"
        elif isinstance(ata_raw, (int, float)):
            # Excel date serial
            try:
                ata_date = (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(ata_raw))).isoformat()
            except Exception:
                pass
            st = status_raw.upper()
            status_type = "LOADING" if "LOADING" in st else "FILED" if "FILED" in st else st
        else:
            # No ATA — check if status contains ETA
            st = status_raw.upper()
            if st.startswith("ETA"):
                status_type = "ETA"
                eta_d = parse_eta_date(status_raw, report_date)
                if eta_d:
                    eta_date = eta_d.isoformat()
            elif "LOADING" in st:
                status_type = "LOADING"
            elif "FILED" in st:
                status_type = "FILED"
            else:
                status_type = st or None

        commodity_code, commodity_name = normalize_commodity(str(commodity_r or "").strip())
        mt_numeric = parse_mt(mt_raw)
        destination = str(dest_raw or "").strip().upper()
        if destination in ("RVT", ""):
            destination = None

        terminal_buffer.append({
            "report_date":     str(report_date),
            "region":          region,
            "terminal_abbrev": current_terminal_abbrev,
            "terminal_name":   current_terminal_name,
            "mile_ahp":        current_mile_ahp,
            "vessel_name":     vessel_name,
            "ata_date":        ata_date,
            "eta_date":        eta_date,
            "status_type":     status_type,
            "mt_raw":          str(mt_raw or "").strip() or None,
            "mt_numeric":      mt_numeric,
            "commodity_code":  commodity_code,
            "commodity_name":  commodity_name,
            "destination":     destination,
            "delay_days_min":  None,   # backfilled when EXPECTED DELAYS row hits
            "delay_days_max":  None,
            "source_file":     source_file,
        })

    _flush_buffer()  # catch any remaining records at end of sheet
    return records


def parse_lineup_file(xlsx_path: Path) -> tuple[list[dict], list[dict]]:
    """
    Parse one SOUTHPORT-LINEUP xlsx. Returns:
      (lineup_records, elevator_ref_records)
    """
    report_date = parse_date_from_filename(xlsx_path.name)
    if report_date is None:
        print(f"    [WARN] Cannot parse date from: {xlsx_path.name} — skipping")
        return [], []

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as e:
        print(f"    [ERROR] Cannot open {xlsx_path.name}: {e}")
        return [], []

    lineup_records = []
    elevator_refs  = []

    sheet_map = {
        "MISS RIVER": ("MISS_RIVER", 0),
        "TEXAS":      ("TEXAS",      0),
        "PNW":        ("PNW",        1),
    }

    for sheet_suffix, (region, col_offset) in sheet_map.items():
        # Match sheet name containing the suffix
        target_sheet = None
        for sname in wb.sheetnames:
            if sheet_suffix in sname.upper() and "SORT" not in sname.upper() and "SAIL" not in sname.upper():
                target_sheet = sname
                break

        if target_sheet is None:
            continue

        ws = wb[target_sheet]
        rows = list(ws.iter_rows(values_only=True))

        if region == "MISS_RIVER":
            elevator_refs = parse_miss_elevator_ref(rows, report_date)

        recs = parse_lineup_sheet(ws, region, col_offset, report_date, xlsx_path.name)
        lineup_records.extend(recs)

    wb.close()
    return lineup_records, elevator_refs


# ── FGIS parser ───────────────────────────────────────────────────────────────

def parse_fgis_file(xlsx_path: Path) -> list[dict]:
    """
    Parse one FGIS Report xlsx → list of cert records from the *DATA sheet.
    Handles both annual (FGIS2022DATA) and weekly (Stats sheet with pivot) formats.
    Uses the DATA sheet which has one row per certificate.
    """
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as e:
        print(f"    [ERROR] Cannot open {xlsx_path.name}: {e}")
        return []

    # Prefer FGISyyyyDATA sheet (detailed, one row per inspection component).
    # Fall back to FGISyyyy (Q4 2025+ summary format — same certs, fewer columns).
    data_sheet = None
    for sname in wb.sheetnames:
        if re.match(r"FGIS\d{4}DATA$", sname.strip(), re.IGNORECASE):
            data_sheet = sname
            break
    if data_sheet is None:
        for sname in wb.sheetnames:
            if re.match(r"FGIS\d{4}$", sname.strip(), re.IGNORECASE):
                data_sheet = sname
                break

    if data_sheet is None:
        wb.close()
        return []

    ws = wb[data_sheet]
    records = []
    header_cols: dict[str, int] = {}

    # Key columns we want (exact match against lowercased header cell)
    # Aliases cover both old format (FGIS2021DATA) and new format (FGIS2025 without DATA)
    WANTED = {
        "serial_no":     ["serial no", "serial no."],
        "cert_date":     ["cert date"],
        "vessel":        ["carrier name", "vessel"],
        "grain":         ["grain"],
        "grain_class":   ["class"],
        "grade":         ["grade"],
        "pounds":        ["pounds"],
        "metric_ton":    ["metric ton", "metric tons"],
        "destination":   ["destination"],
        "port":          ["port"],
        "field_office":  ["field office"],
        "ams_region":    ["ams reg"],
        "fgis_region":   ["fgis reg"],
        "moisture_avg":  ["m avg", "moisture avg", "moisture"],
        "fm_pct":        ["fm", "foriegn matter", "foreign matter", "fm%"],
        "protein_avg":   ["p avg", "protein avg", "protein"],
        "oil_avg":       ["o avg", "oil avg", "oil"],
        "test_weight":   ["tw", "test wt", "test weight", "tw avg"],
    }

    header_found = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not header_found:
            # Scan dynamically for the header row (may be row 0 or offset by title block)
            candidate: dict[str, int] = {}
            for j, cell in enumerate(row):
                label = str(cell or "").strip().lower()
                for field, aliases in WANTED.items():
                    if label in aliases and field not in candidate:
                        candidate[field] = j
            if len(candidate) >= 3:   # found header: at least 3 recognized columns
                header_cols = candidate
                header_found = True
            continue

        if not row or all(v is None for v in row):
            continue

        def g(field: str):
            idx = header_cols.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        # Parse cert_date — could be datetime, Excel serial (>30000), or YYYYMMDD int (8 digits)
        cert_date_raw = g("cert_date")
        cert_date: Optional[str] = None
        if isinstance(cert_date_raw, datetime.datetime):
            cert_date = cert_date_raw.date().isoformat()
        elif isinstance(cert_date_raw, (int, float)):
            raw_int = int(cert_date_raw)
            if raw_int >= 20000101 and raw_int <= 20991231:
                # YYYYMMDD format
                try:
                    s = str(raw_int)
                    cert_date = f"{s[:4]}-{s[4:6]}-{s[6:]}"
                except Exception:
                    cert_date = None
            elif raw_int > 30000:
                # Excel date serial
                try:
                    cert_date = (datetime.date(1899, 12, 30) + datetime.timedelta(days=raw_int)).isoformat()
                except Exception:
                    cert_date = None

        # Parse pounds
        pounds_raw = g("pounds")
        try:
            pounds = float(pounds_raw) if pounds_raw is not None else None
        except (TypeError, ValueError):
            pounds = None

        # Parse metric tons — could be in pounds column recalculated or dedicated column
        mt_raw = g("metric_ton")
        try:
            mt = float(mt_raw) if mt_raw is not None else (pounds / 2204.62 if pounds else None)
        except (TypeError, ValueError):
            mt = None

        serial_no = g("serial_no")
        if serial_no is None:
            continue

        records.append({
            "serial_no":    str(serial_no).strip(),
            "cert_date":    cert_date,
            "vessel":       str(g("vessel") or "").strip() or None,
            "grain":        str(g("grain") or "").strip() or None,
            "grain_class":  str(g("grain_class") or "").strip() or None,
            "grade":        str(g("grade") or "").strip() or None,
            "pounds":       pounds,
            "metric_ton":   mt,
            "destination":  str(g("destination") or "").strip() or None,
            "port":         str(g("port") or "").strip() or None,
            "field_office": str(g("field_office") or "").strip() or None,
            "ams_region":   str(g("ams_region") or "").strip() or None,
            "fgis_region":  str(g("fgis_region") or "").strip() or None,
            "moisture_avg": _safe_float(g("moisture_avg")),
            "fm_pct":       _safe_float(g("fm_pct")),
            "protein_avg":  _safe_float(g("protein_avg")),
            "oil_avg":      _safe_float(g("oil_avg")),
            "test_weight":  _safe_float(g("test_weight")),
            "source_file":  xlsx_path.name,
        })

    wb.close()
    return records


def _safe_float(v) -> Optional[float]:
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


# ── DuckDB helpers ────────────────────────────────────────────────────────────

def create_tables(con: duckdb.DuckDBPyConnection) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS grain_southport_lineup (
            report_date       DATE,
            region            VARCHAR,
            terminal_abbrev   VARCHAR,
            terminal_name     VARCHAR,
            mile_ahp          DOUBLE,
            vessel_name       VARCHAR,
            ata_date          DATE,
            eta_date          DATE,
            status_type       VARCHAR,
            mt_raw            VARCHAR,
            mt_numeric        DOUBLE,
            commodity_code    VARCHAR,
            commodity_name    VARCHAR,
            destination       VARCHAR,
            delay_days_min    INTEGER,
            delay_days_max    INTEGER,
            source_file       VARCHAR
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS grain_miss_elevator_ref (
            report_date       DATE,
            mile_ahp          DOUBLE,
            elevator_name     VARCHAR,
            air_height_spouts VARCHAR,
            gauge_reference   VARCHAR
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS grain_fgis_certs (
            serial_no     VARCHAR,
            cert_date     DATE,
            vessel        VARCHAR,
            grain         VARCHAR,
            grain_class   VARCHAR,
            grade         VARCHAR,
            pounds        DOUBLE,
            metric_ton    DOUBLE,
            destination   VARCHAR,
            port          VARCHAR,
            field_office  VARCHAR,
            ams_region    VARCHAR,
            fgis_region   VARCHAR,
            moisture_avg  DOUBLE,
            fm_pct        DOUBLE,
            protein_avg   DOUBLE,
            oil_avg       DOUBLE,
            test_weight   DOUBLE,
            source_file   VARCHAR
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS grain_ingest_log (
            filename      VARCHAR PRIMARY KEY,
            table_name    VARCHAR,
            rows_loaded   INTEGER,
            ingested_at   TIMESTAMP DEFAULT now()
        )
    """)


def already_ingested(con: duckdb.DuckDBPyConnection, filename: str) -> bool:
    r = con.execute(
        "SELECT COUNT(*) FROM grain_ingest_log WHERE filename = ?", [filename]
    ).fetchone()
    return r[0] > 0


def log_ingestion(con: duckdb.DuckDBPyConnection, filename: str, table: str, rows: int) -> None:
    con.execute(
        """INSERT OR REPLACE INTO grain_ingest_log (filename, table_name, rows_loaded, ingested_at)
           VALUES (?, ?, ?, now())""",
        [filename, table, rows]
    )


def bulk_insert(con: duckdb.DuckDBPyConnection, table: str, records: list[dict]) -> None:
    if not records:
        return
    cols = list(records[0].keys())
    placeholders = ", ".join(["?"] * len(cols))
    col_str = ", ".join(cols)
    stmt = f"INSERT INTO {table} ({col_str}) VALUES ({placeholders})"
    values = [[r[c] for c in cols] for r in records]
    con.executemany(stmt, values)


# ── Main pipeline ─────────────────────────────────────────────────────────────

def run_lineup_ingest(con: duckdb.DuckDBPyConnection, force: bool) -> dict:
    files = sorted(SOUTHPORT_DIR.glob("*LINEUP*.xlsx")) + sorted(SOUTHPORT_DIR.glob("*Lineup*.xlsx"))
    # Deduplicate
    files = list(dict.fromkeys(files))
    print(f"\nSOUTHPORT-LINEUP files found: {len(files)}")

    processed = skipped = failed = 0
    total_rows = 0

    for xlsx_path in files:
        fname = xlsx_path.name
        if not force and already_ingested(con, fname):
            skipped += 1
            continue

        print(f"  {fname}", end="", flush=True)
        lineup_recs, elevator_recs = parse_lineup_file(xlsx_path)

        if not lineup_recs and not elevator_recs:
            print(f" → [SKIP: no data]")
            failed += 1
            continue

        # Remove existing rows for this file before re-insert (idempotent)
        con.execute("DELETE FROM grain_southport_lineup WHERE source_file = ?", [fname])
        con.execute("DELETE FROM grain_miss_elevator_ref WHERE report_date = (SELECT report_date FROM grain_southport_lineup WHERE source_file = ? LIMIT 1)", [fname])

        bulk_insert(con, "grain_southport_lineup", lineup_recs)
        bulk_insert(con, "grain_miss_elevator_ref", elevator_recs)

        total_rows += len(lineup_recs)
        log_ingestion(con, fname, "grain_southport_lineup", len(lineup_recs))
        processed += 1
        print(f" → {len(lineup_recs)} vessel rows")

    return {"processed": processed, "skipped": skipped, "failed": failed, "total_rows": total_rows}


def run_fgis_ingest(con: duckdb.DuckDBPyConnection, force: bool) -> dict:
    files = sorted(SOUTHPORT_DIR.glob("FGIS*.xlsx")) + sorted(SOUTHPORT_DIR.glob("FFGIS*.xlsx"))
    files = list(dict.fromkeys(files))
    print(f"\nFGIS files found: {len(files)}")

    processed = skipped = failed = 0
    total_rows = 0

    for xlsx_path in files:
        fname = xlsx_path.name
        if not force and already_ingested(con, fname):
            skipped += 1
            continue

        print(f"  {fname}", end="", flush=True)
        recs = parse_fgis_file(xlsx_path)

        if not recs:
            # Many weekly FGIS files only have pivot Stats sheet, not DATA sheet — normal
            print(f" → [no DATA sheet — skip]")
            log_ingestion(con, fname, "grain_fgis_certs", 0)
            continue

        con.execute("DELETE FROM grain_fgis_certs WHERE source_file = ?", [fname])
        bulk_insert(con, "grain_fgis_certs", recs)
        total_rows += len(recs)
        log_ingestion(con, fname, "grain_fgis_certs", len(recs))
        processed += 1
        print(f" → {len(recs):,} cert rows")

    return {"processed": processed, "skipped": skipped, "failed": failed, "total_rows": total_rows}


def print_summary(con: duckdb.DuckDBPyConnection) -> None:
    print("\n" + "=" * 60)
    print("Table Summary")
    print("=" * 60)
    for table in ["grain_southport_lineup", "grain_miss_elevator_ref", "grain_fgis_certs", "grain_ingest_log"]:
        try:
            r = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()
            print(f"  {table:<35} {r[0]:>10,} rows")
        except Exception:
            print(f"  {table:<35}  [not found]")

    # Quick stats
    print("\nLineup coverage:")
    r = con.execute("""
        SELECT MIN(report_date), MAX(report_date), COUNT(DISTINCT report_date),
               COUNT(DISTINCT vessel_name), SUM(mt_numeric)/1e6
        FROM grain_southport_lineup
    """).fetchone()
    if r and r[0]:
        print(f"  Date range:      {r[0]} → {r[1]}")
        print(f"  Report days:     {r[2]:,}")
        print(f"  Unique vessels:  {r[3]:,}")
        print(f"  Total cargo (M MT): {r[4]:.1f}" if r[4] else "  Total cargo:    (calculating...)")

    print("\nTop destinations (lineup, by MT):")
    rows = con.execute("""
        SELECT destination, ROUND(SUM(mt_numeric)/1000) as kt
        FROM grain_southport_lineup
        WHERE destination IS NOT NULL AND mt_numeric IS NOT NULL
        GROUP BY 1 ORDER BY 2 DESC LIMIT 10
    """).fetchall()
    for dest, kt in (rows or []):
        print(f"  {dest:<25} {kt:>8,.0f} KT")


def main():
    parser = argparse.ArgumentParser(description="Southport Agencies Excel → DuckDB ingestion")
    parser.add_argument("--lineup",  action="store_true", help="Process SOUTHPORT-LINEUP files only")
    parser.add_argument("--fgis",    action="store_true", help="Process FGIS Report files only")
    parser.add_argument("--force",   action="store_true", help="Reprocess already-ingested files")
    parser.add_argument("--summary", action="store_true", help="Print table summary and exit")
    args = parser.parse_args()

    do_lineup = args.lineup or (not args.lineup and not args.fgis)
    do_fgis   = args.fgis   or (not args.lineup and not args.fgis)

    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect(str(DB_PATH))
    create_tables(con)

    if args.summary:
        print_summary(con)
        con.close()
        return

    print("=" * 60)
    print("Southport Agencies — Excel Ingestion Pipeline")
    print("=" * 60)
    print(f"Source dir:  {SOUTHPORT_DIR}")
    print(f"Database:    {DB_PATH}")
    print(f"Force mode:  {args.force}")

    if not SOUTHPORT_DIR.exists():
        print(f"\n[ERROR] Source directory not found: {SOUTHPORT_DIR}")
        sys.exit(1)

    if do_lineup:
        stats = run_lineup_ingest(con, args.force)
        print(f"\nLINEUP: {stats['processed']} processed, {stats['skipped']} skipped, "
              f"{stats['failed']} failed, {stats['total_rows']:,} rows")

    if do_fgis:
        stats = run_fgis_ingest(con, args.force)
        print(f"\nFGIS:   {stats['processed']} processed, {stats['skipped']} skipped, "
              f"{stats['failed']} failed, {stats['total_rows']:,} rows")

    con.close()
    print("\nDone. Run with --summary to see table statistics.")


if __name__ == "__main__":
    main()
